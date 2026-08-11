from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from nanobot.rag.config import RagParsingConfig
from nanobot.rag.parser import ParseCompleteness, RagParseError, parse_document
from nanobot.rag.types import RagErrorCode, SourceKind


def test_xlsx_preserves_sheet_and_one_based_row_ranges(tmp_path: Path) -> None:
    path = tmp_path / "inventory.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存"
    sheet.append(["名称", "数量"])
    sheet.append(["苹果", 3])
    second = workbook.create_sheet("说明")
    second.append(["仅供测试"])
    workbook.save(path)

    result = parse_document(path, RagParsingConfig())

    assert [block.text for block in result.blocks] == ["名称\t数量", "苹果\t3", "仅供测试"]
    assert all(block.location.kind is SourceKind.SPREADSHEET_ROWS for block in result.blocks)
    assert result.blocks[1].location.sheet == "库存"
    assert result.blocks[1].location.row_start == 2
    assert result.blocks[1].location.row_end == 2
    assert result.metrics.table_rows == 3
    assert result.metrics.table_cells == 5


def test_xlsx_reads_display_cache_only_and_never_formula_source(tmp_path: Path) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "visible"
    sheet["A2"] = "=HYPERLINK(\"https://example.invalid\",\"remote\")"
    workbook.save(path)

    result = parse_document(path, RagParsingConfig())

    assert [block.text for block in result.blocks] == ["visible"]
    assert "HYPERLINK" not in "\n".join(block.text for block in result.blocks)


def test_xlsx_row_limit_is_a_safe_error(tmp_path: Path) -> None:
    path = tmp_path / "rows.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["one"])
    sheet.append(["two"])
    workbook.save(path)

    with pytest.raises(RagParseError) as exc_info:
        parse_document(path, RagParsingConfig(max_table_rows=1))

    assert exc_info.value.code is RagErrorCode.UNSAFE_DOCUMENT


def test_xlsx_character_limit_is_explicit(tmp_path: Path) -> None:
    path = tmp_path / "large.xlsx"
    workbook = Workbook()
    workbook.active.append(["0123456789", "abcdefghij"])
    workbook.save(path)

    result = parse_document(path, RagParsingConfig(max_extracted_chars=8))

    assert result.completeness is ParseCompleteness.TRUNCATED
    assert result.limit_reason == "max_extracted_chars"
    assert result.total_chars <= 8


def test_empty_xlsx_has_no_extractable_text(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)

    with pytest.raises(RagParseError) as exc_info:
        parse_document(path, RagParsingConfig())

    assert exc_info.value.code is RagErrorCode.NO_EXTRACTABLE_TEXT
